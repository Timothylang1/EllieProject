from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Side, Border
from openpyxl.utils import get_column_letter


def createExcel(filepath, titleOfSheet, data, fieldnames, fillIn):
    # Slight check to ensure that filepath has ".xlsx" at the end
    if filepath[len(filepath) - 5: len(filepath)] != ".xlsx":
        filepath += ".xlsx"

    try:
        # Attempts to open a workbook at that location
        wb = load_workbook(filepath)
        wb.create_sheet(titleOfSheet) # Creates a new sheet
        ws = wb.worksheets[len(wb.worksheets) - 1] # Gets the last sheet
        addDataIn(ws, data, fieldnames)
        if fillIn: # If the user wants to use the fill in feature
            matches = findMatches(wb, fieldnames, data) # Returns a list with the number of matches each row had with the previous worksheets
            colorIn(ws, matches)
    except:
        # Otherwise, it creates a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = titleOfSheet
        addDataIn(ws, data, fieldnames)
    
    # Saves the worksheet at the filepath
    wb.save(filepath)


def addDataIn(worksheet, data, fieldnames):
    """Adds in data to the worksheet using the fieldnames as headers."""
    # Append headers
    worksheet.append(fieldnames)
    dim = []
    for field in fieldnames:
        dim.append(len(field)) # Used for find the maximum dimension needed for column to fit all data. Width starts with length of headers
    for row in data:
        toAdd = []
        for fieldnum in range(len(fieldnames)):
            value = row[fieldnames[fieldnum]]
            dim[fieldnum] = max(dim[fieldnum], len(value)) # Updates column width tracker if we need to adjust the width of the column to fit the new data
            columnValue = testFloat(value)
            toAdd.append(columnValue)
        worksheet.append(toAdd) # Adds in the row of now ordered values
    for colWidth in range(len(dim)):
        worksheet.column_dimensions[get_column_letter(colWidth + 1)].width = dim[colWidth] + 1 # Columns start at 1, and we increase the width by 1 so that data isn't crammed in


def findMatches(workbook, fieldnames, data):
    """Looks through past sheets, and if any of the columns info matches with the last sheet, it highlights the last sheet row of data.
    The more prevelant the data is, the darker the color. Doesn't affect previous sheets"""
    # Creates a list that tracks how many matches for each row
    matchesPerRow = [0] * len(data)

    for worksheet in workbook.worksheets[0:len(workbook.worksheets) - 1]: # Not including the last one since we already have that info
        headers = worksheet[1] # Gets the first row of data containing all of the headers
        matchingColumnHeaders = {} # Keeps track of which columns have matching headers with headers in the new dataset
        for header in headers: # For each header in this worksheet
            if header.value in fieldnames: # If the header is one of the values we want to use for comparison, then we iterate down that column
                matchingColumnHeaders[header.value] = header.column - 1 # Gets the column where column 1 is stored as 0
        
        # Removes columns we don't need for comparison from the original data set
        tempData = []
        for row in data:
            newRow = {} # Creates new dictionary to add to tempData row containing only the elements we need
            for header in row:
                if header in matchingColumnHeaders:
                    newRow[header] = row[header]
            tempData.append(newRow)

        # Now we iterate over each row, and check if its data set matches
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            # Create temporary dictionary of column names : data in row
            temp = {}
            for header in matchingColumnHeaders:
                temp[header] = str(row[matchingColumnHeaders[header]]) # Converts whatever data to string so that we can test if the two strings are the same

            # See if dictionary is in the list, and update it's corresponding counter.
            for dataRow in range(len(tempData)):
                if tempData[dataRow] == temp: # Compares the two dictionaries
                    matchesPerRow[dataRow] += 1

    # After we know how many matches there are per row in all of the spreadsheets, we can fill in the correct rows with the color based on the number of matches
    # The more matches, the darker the color so that it stands out more
    return matchesPerRow


def colorIn(worksheet, matches):
    """Takes in the current worksheet, and colors in each row based on how many matches it had with the previous spreadsheets.
    The more matches, the darker the color."""
    if max(matches) != 0: # If there were any matches at all...
        rowTracker = 0
        colorDifferenceG = 255 / max(matches)
        colorDifferenceRB = 50 / max(matches)
        for row in worksheet.iter_rows(min_row=2):
            # Calculating rgb value as a hexidecimal. The more matches, the darker it becomes. Also formats the g value
            color = "{:02x}".format(255 - round(colorDifferenceRB * matches[rowTracker])) + "{:02x}".format(255 - round(colorDifferenceG * matches[rowTracker])) + "{:02x}".format(255 - round(colorDifferenceRB * matches[rowTracker]))
            fill_cell = PatternFill(patternType='solid', fgColor=color)

            # Then changes all the colors
            for cell in row:
                cell.fill = fill_cell
                thin = Side(border_style="thin", color="C0C0C0")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            rowTracker += 1


def testFloat(test):
    try:
        return float(test)
    except ValueError:
        return test

# Testing

# testDict = [{"name": "Timothy", "address": "1095 McGregor Way"},
#             {"name": "Paul", "address": "3707 La Calle Ct"},
#             {"name": "Alice", "address": "790 La Para Ave"},
#             {"name": "Excel", "address": "3953 Laguna Ave"}]

# fieldnamesForTest = ["name", "address"]

# filepathTest = "sample.xlsx"

# titleForTest = "test"

# fillIn = True

# createExcel(filepathTest, titleForTest, testDict, fieldnamesForTest, fillIn)
